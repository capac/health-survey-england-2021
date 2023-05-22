from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

plt.style.use('barplot-style.mplstyle')

# data directory
data_dir = Path.home() / 'Programming/data/health-survey-england-2021/'

# working directory
python_work_dir = Path.home() / 'Programming/Python/'
work_dir = python_work_dir / 'data-visualization/health-survey-england-2021/'

workbook = load_workbook(filename=data_dir /
                         'HSE-2021-Overweight-and-obesity-tables.xlsx')

wb_table1 = workbook['Table 1']

weight_dict = {}
age_list = ['16-24', '25-34', '35-44', '45-54',
            '55-64', '65-74', '75+',]
weight_list = ['Neither obese or overweight', 'Overweight', 'Obese',]

for cat, row in zip(weight_list, wb_table1.iter_rows(min_row=24, max_row=26,
                                                     min_col=2, max_col=8,
                                                     values_only=True)):
    weight_dict[cat] = row

# stacked bar plot
fig, axes = plt.subplots()
bottom = np.zeros(7)
for label, weight_per in weight_dict.items():
    p = axes.bar(age_list, weight_per, bottom=bottom,
                 label=label, alpha=0.95, width=0.7)
    bottom += weight_per

bottom_values = np.ones(7)
for values in weight_dict.values():
    for label, value, b_value in zip(age_list, values, bottom_values):
        axes.annotate('{:g} %'.format(round(value, 0)),
                      (label, b_value+value/2), fontsize=14,
                      fontweight='bold', ha='center', va='top')
    bottom_values += values

axes.legend(loc='lower center', fontsize=12, fancybox=True,
            bbox_to_anchor=(0.5, -0.13), ncols=3)
axes.set_xlabel('Age Range')
axes.set_ylabel('Percentage (%)')
axes.set_title('Percentage of neither obese or overweight, overweight and obese in England (2021)')

plt.savefig(work_dir / 'plots/weight_charts_2021.png')
